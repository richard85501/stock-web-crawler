from bs4 import BeautifulSoup
import requests
import time
import random
import openpyxl
from openpyxl import load_workbook
# 修改目前工作表
# from openpyxl.utils import get_column_letter, column_index_from_string

##-------------------------- openpyxl----------------------------##

fn = 'test.xlsx'
# 預設可讀寫，若有需要可以指定write_only和read_only為True
wb = load_workbook(fn)

# 目前在第幾個工作表,索引值從0開始
wb.active = 0
print('excel活動工作表： ', wb.active)
ws = wb.active

# 記得開vpn
# 連結網站

##
headers = {'user-agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36'}
response = requests.get(
	"https://histock.tw/stock/brokerprofit.aspx?bno=1470", headers=headers)

# //理論上應該要設置為隨機
# time.sleep(20)
# HTML原始碼解析
bsObj = BeautifulSoup(response.text, "html.parser")

# b是一列的變數
b=1 
# a = bsObj.find("table", {"id" : "CPHB1_bt1_g"})
# print(a)
for title in bsObj.find("table", {"id" : "CPHB1_bt1_g"}).find("tr"):
    # a += title.string;
    print(title.string)
    #寫入儲存格
    ws.cell(column=b, row=1).value = title.string
    b+=1
# print(a)
##

# rate0A = 2

# for single_tr in bsObj.find("table", {"id" : "CPHB1_bt1_g"}).findAll("tr"):
# 	cell = single_tr.findAll("td")
# 	if len(cell) != 0 :
# 		# print(cell)
# 		rate0 = cell[0].find('a').string
# 		rate1 = cell[1].contents[0]
# 		rate2 = cell[2].contents[0]
# 		rate3 = cell[3].contents[0]
# 		rate4 = cell[4].contents[0]
# 		rate5 = cell[5].contents[0]
# 		rate6 = cell[6].contents[0]
# 		rate7 = cell[7].contents[0]
# 		rate8 = cell[8].contents[0]
# 		rate9 = cell[9].contents[0]
# 		rate10 = cell[10].contents[0]
# 		rate11 = cell[11].contents[0]

# 		print(rate0)

# 		ws.cell(column=2, row=rate0A).value = rate0
# 		ws.cell(column=3, row=rate0A).value = rate1
# 		ws.cell(column=4, row=rate0A).value = rate2
# 		ws.cell(column=5, row=rate0A).value = rate3
# 		ws.cell(column=6, row=rate0A).value = rate4
# 		ws.cell(column=7, row=rate0A).value = rate5
# 		ws.cell(column=8, row=rate0A).value = rate6
# 		ws.cell(column=9, row=rate0A).value = rate7
# 		ws.cell(column=10, row=rate0A).value = rate8
# 		ws.cell(column=11, row=rate0A).value = rate9
# 		ws.cell(column=12, row=rate0A).value = rate10
# 		ws.cell(column=13, row=rate0A).value = rate11
# 		rate0A += 1


# 儲存檔案(必須把excel檔關掉始可執行)
wb.save(fn)